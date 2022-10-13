from openpyxl import load_workbook
from Framework import framework
import datetime
import time

print('Enter the Excel file location file ')
#path = 'D:\ReadData\ExcelData.xlsx'
path = input()
try:
  wb = load_workbook(path)
  sh = wb['Test']
except Exception as e:
  print('The Excel file is not found',e)
  time.sleep(4)
print('Enter the location of the chrome driver ')
ChromePATH = input()
#ChromePATH = "D:\ReadData\chromedriver.exe"
print('Enter the starting point')
startat=int(input())
print('Enter the ending point')
endat=int(input())
Vacation = 21
NoticePeriod=60
Probation=90
Annualleave =21
Workingdays=6
workhours= 48
workdomain = 'Indoor'
worktype = 'Full Time'
bankaccount = 2
ContractPeriod='weekly'
workperiod = 'Fixed Contract'
try:
 frame = framework(ChromePATH)
except Exception as E:
    print('Check chrome driver ',E)
print('Processing....')
start = time.time()
for i in range(startat,endat+1):
    EmployeeEnglishName = sh.cell(i,2).value
    Nationality=sh.cell(i,5).value
    City=sh.cell(i,11).value
    JobEnglish= sh.cell(i,7).value
    JobArabic = sh.cell(i,6).value
    Religion= sh.cell(i,12).value
    dates = sh.cell(i,13).value
    EndDate=sh.cell(i,14).value
    #ContractEnd = datetime.date.strftime(EndDate, '%Y-%m-%d')
    ContractEnd = EndDate
    ContractStart = datetime.date.strftime(dates, '%Y-%m-%d')
    BankCard=sh.cell(i,15).value
    BankState = sh.cell(i,16).value
    BankName=sh.cell(i,17).value
    Iqama = sh.cell(i,9).value
    Annualleave = str(sh.cell(i,18).value)
    try:
     frame.Searchemp(Iqama)
     if frame.checknodata() > 0:
       sh.cell(i,26).value = 0
       continue
     if frame.checkauthenticated()>0:
       #sh.cell(i,26).value=3
       continue
     frame.authenticate()
     if frame.checkabsher() >0:
       frame.handleabsher()
       sh.cell(i,26).value =2
       continue
     frame.chooseReligion(Religion)
     frame.EnterjobTitle(JobEnglish,JobArabic)
     frame.chooseWorktype(worktype)
     frame.chooseCity(City)
     #frame.chooseWorkDomain(workdomain)
     #if workdomain == 'Abroad':
     #frame.chooseCountry()
     #else:
    #frame.chooseCity(City)
     if BankState == '2':
      frame.chooseSalarycard()
      frame.writeIBan(BankCard)
      frame.writeBank(BankName)
     else:
      frame.writeIBan(BankCard)
     time.sleep(1)
     frame.Next()
     frame.chooseContractPeriod(workperiod)
     frame.writeProbation(Probation)
     frame.writeDate(ContractStart)
     frame.writeEndPeriod(ContractEnd)
     frame.chooseWorkingHours(ContractPeriod)
     frame.writeNoticeperiod(NoticePeriod)
     frame.writeAnualleave(Annualleave)
     frame.writeWeeklyworkingDays(Workingdays)
     frame.writeWorkingHours(workhours)
     frame.Nextv2()
     time.sleep(2)
     count = frame.checkadditional()
     while(count<3):
      frame.Previus()
      frame.Nextv2()
      count = frame.checkadditional()
     frame.agreeAdditionalterm()
     frame.submit()
     sh.cell(i, 26).value = 1
     frame.waitLoading()
    except Exception as s:
        wb.save(path)
        print('Error Happens at Line ',i)
        print('The error was ',s)
        time.sleep(20)
end = time.time()
hours, rem = divmod(end-start, 3600)
minutes, seconds = divmod(rem, 60)
print("{:0>2}:{:0>2}:{:05.2f}".format(int(hours),int(minutes),seconds))
wb.save(path)
#frame.close()






